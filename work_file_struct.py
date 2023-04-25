# обработка файлов Excel из директории D:\Каштан\Файлы_23
# выгрузка продукции без выделения номенклатуры
# выбор файла через оконное приложение
#
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, PatternFill, \
            Font,  Alignment

from tkinter import *
from tkinter import Frame, Tk, BOTH, W, N, E, S,  Text, Button, END
from tkinter import filedialog

class Example(Frame):

    def __init__(self):
        super().__init__()
        self.initUI()

    def initUI(self):
        self.master.title("Окно для выбора файла")
        self.pack(fill=BOTH, expand=1)

        select = Button(self, text="Выбрать файл", command=self.onOpen)
        select.grid(row=0, column=0, padx=5, pady=5, sticky=N)


        self.txt = Text(self)
        self.txt.grid(row=1, column=1)

    def onOpen(self):
        lbl = Label(self, text=" ")
        lbl.grid(row=1, column=0, padx=5, pady=5, sticky=N)

        ftypes = [('Python файлы', '*.py'), ('Все файлы', '*')]
        dlg = filedialog.Open(self, filetypes=ftypes)
        fl = dlg.show()
        print("fl=", fl)
        self.txt.delete("1.0", END)
        if fl != '':
            text = self.readFile(fl)
            self.txt.insert(END, text)
            lbl["text"] = "Закрыть х"
            self.work_file(fl)

    def readFile(self, filename):
        text = filename
        return text

    def work_file(self, fl):
        wb = load_workbook(fl)
        ws1 = wb['Лист1']
        ws2 = wb['Лист2']
        print(wb.sheetnames)
        print(f"размер листа {ws1} = {ws1.dimensions}")

        # заполняем '0' ячейки с None или ''
        for row in ws1.iter_rows(min_row=6, min_col=6, max_col=ws1.max_column, max_row=ws1.max_row):

            for cell in row:
                # print("cell.value=", cell.value, type(cell.value))
                if cell.value == ' ' or cell.value is None:
                    cell.value = 0
                    # print("cell.value  ' '", cell.value, type(cell.value))
        # выбираем по всей таблице  сначения строк в столбце А
        lst_row = []
        for row in ws1.iter_rows(min_row=6, min_col=1, max_col=1, max_row=ws1.max_row, values_only=True):
            if row[0]:
                lst_row.append(row)

        print(f"количество наименований УК в листе {wb.sheetnames[0]} = {len(lst_row)} ")
        print(f"количество столбцов  в {ws1} = {ws1.max_column}")
        print(f"количество строк в {ws1} = {ws1.max_row}")

        # изменяем ширину колонок (столбцов)
        ws2.column_dimensions['B'].width = 35
        ws2.column_dimensions['C'].width = 10
        ws2.column_dimensions['M'].width = 15
        # изменяем высоту строк
        ws2.row_dimensions[5].height = 105

        # выбираем список первых 5  строк таблицы Лист1
        # передаем значения ячеек первых пяти строк  из первой таблицы во вторую
        for i in range(1, 6):
            for j in range(1, ws1.max_column + 1):
                ws2.cell(row=i, column=j).value = ws1.cell(row=i, column=j).value
        print(f"передача значения ячеек первых 5 строк  {ws1} таблицы во {ws2}")

        # определим стили границ
        mediums = Side(border_style="medium", color="000000")
        # выравнивание по центру ячейки
        for cell in ws2[5]:
            if cell.value:
                cell.alignment = Alignment(horizontal='center', vertical='center',
                                           text_rotation=0,
                                           wrap_text=True,
                                           shrink_to_fit=False,
                                           indent=0)
                cell.border = Border(top=mediums, left=mediums, right=mediums, bottom=mediums)
                cell.font = Font(bold=False, color='000000', name='Arial', size=8)
        print("выравнивание по центру ячейки 5 строки заверщено")

        # прейскурантная цена с 06 10 2022 row[6]
        price_prev = 0
        # примеменяемая цена по 05 10 2022 row[7]
        primen_prev = 0
        # отпускная цена с 06 10 2022 row[8]
        otpuckn_prev = 0

        # выбираем по всей таблице  значения строк
        list_row = []
        price = []
        primen = []
        otpuckn = []
        for row in ws1.iter_rows(min_row=6, min_col=1, max_col=15, max_row=ws1.max_row, values_only=True):
            # выбираем по всей таблице  значение ячейки строки в столбце А
            # print("row[0]=", row[0], type(row[0]))
            if row[0] and row[0] != ' ':
                print(f"прейскурантная цена= {price}; применяемая цена= {primen};"
                      f" отпускная цена= {otpuckn};")
                price_prev = 0
                primen_prev = 0
                otpuckn_prev = 0
                ws2.append(row)
                print(row[0])
                price = []
                primen = []
                otpuckn = []
                # list_row.append(row[0])
            else:
                price_cur = row[6]
                primen_cur = row[7]
                otpuckn_cur = row[8]
                counter = 0
                if price_cur != price_prev:
                    price_prev = price_cur
                    ws2.append(row)
                    counter += 1
                    price.append(price_prev)
                if primen_cur != primen_prev:
                    primen_prev = primen_cur
                    primen.append(primen_prev)
                    if counter == 0:
                        ws2.append(row)
                        counter += 1
                if otpuckn_cur != otpuckn_prev:
                    otpuckn_prev = otpuckn_cur
                    otpuckn.append(otpuckn_prev)
                    if counter == 0:
                        ws2.append(row)

        print(f"выбирали по всей таблице {ws1} значения ячеек строк и переносили в {ws2}")

        thins = Side(border_style="thin", color="000000")
        for i in range(6, ws2.max_row):
            for j in range(1, ws2.max_column + 1):
                # присвоили стиль каждой ячейке в 6 строке и далее
                cell = ws2.cell(row=i, column=j)
                cell.alignment = Alignment(horizontal='left', vertical='center',
                                           text_rotation=0,
                                           wrap_text=True,
                                           shrink_to_fit=False,
                                           indent=0)
                cell.border = Border(top=thins, left=thins, right=thins, bottom=thins)
                cell.font = Font(bold=False, color='000000', name='Arial', size=8)
                # красный цвет текста ячейки,  шрифт name='Arial', size=8
                # cell.font = Font(bold=True, color='FF0000', name='Arial', size=8)
        print(f"форматирование строк с 6 по конец листа {ws2} завершено")

        # выбираем по всей таблице  сначения строк в столбце А
        lst_row2 = []
        for row in ws2.iter_rows(min_row=6, min_col=1, max_col=15, max_row=ws2.max_row, values_only=True):
            if row[0]:
                lst_row2.append(row)
        print()
        print(f"количество наименований УК в листе {wb.sheetnames[1]} = {len(lst_row2)} ")
        print(f"количество столбцов  в {ws2} = {ws2.max_column}")
        print(f"количество строк в {ws2} = {ws2.max_row}")

        for i in range(6, ws2.max_row + 1):
            cell_row = ws2.cell(row=i, column=1).value
            if cell_row and cell_row != ' ':
                ws2.merge_cells(start_row=i, start_column=1, end_row=i, end_column=2)
                cell = ws2.cell(row=i, column=1)
                cell.alignment = Alignment(horizontal='left', vertical='center')
                cell.fill = PatternFill(fill_type='solid', fgColor='FFFF00')
                for j in range(3, ws2.max_column + 1):
                    cell = ws2.cell(row=i, column=j)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.fill = PatternFill(fill_type='solid', fgColor='FFFF00')
        print(f"слияние строке УК столбцов А и В листа {ws2} завершено")

        # выделение красным цветом отгрузок без прейскурантной цены
        for i in range(6, ws2.max_row + 1):
            # присвоили стиль каждой ячейке в 6 строке и далее
            cell7 = ws2.cell(row=i, column=7)
            cell9 = ws2.cell(row=i, column=9)
            if cell9.value != 0 and cell7.value == 0:
                for j in range(2, ws2.max_column + 1):
                    cell = ws2.cell(row=i, column=j)
                    cell.font = Font(bold=False, color='FF0000', name='Arial', size=8)
        print(f"выделение красным цветом отгрузок без прейскурантной цены листа {ws2} завершено")

        wb.save(fl)

def main():
    root = Tk()
    ex = Example()
    root.geometry("400x250+300+300")
    root.mainloop()


if __name__ == '__main__':
    main()
